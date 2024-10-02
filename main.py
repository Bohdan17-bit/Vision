from openpyxl import load_workbook
from model import Model
from form import Ui_MainWindow
from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtWidgets import QFileDialog
import os.path
import requests


def find_freq_for_word(sheet, column, word):
    for cell in column:
        if word == cell.value:
            return sheet.cell(row=cell.row, column=21).value
    return "not found!"


def show_time_to_read(time, sd_time):
    min = time // 60000
    sec = (time % 60000) // 1000
    ms = time % 1000

    min = int(min)
    sec = int(sec)

    min_sd = sd_time // 60000
    sec_sd = (sd_time % 60000) // 1000
    ms_sd = sd_time % 1000

    min_sd = int(min_sd)
    sec_sd = int(sec_sd)

    if min > 0:
        final_result = f"Необхідно: {min},{int(sec_sd) % 100} хв. для читання."
    else:
        final_result = f"Необхідно: {sec},{int(ms) % 100} с. для читання."

    if min_sd > 0:
        path = f"\nСтандартне відхилення складає: {min_sd},{sec_sd} хв."
    else:
        path = f"\nСтандартне відхилення складає: {sec_sd},{int(ms_sd) % 100} с."

    final_result += path

    showResult(final_result)


def showResult(message):
    ui.TotalTime_lineEdit.setText(message)


def get_frequency_dictionary(path):
    return load_workbook(path)


def get_sheet_workbook(workbook, sheet):
    list_sheets = workbook.sheetnames
    if sheet in list_sheets:
        return workbook[sheet]
    else:
        return None


def get_biggest_frequency(sheet):
    return sheet.cell(row=2, column=21).value


def get_column_by_sheet(sheet, name):
    return sheet[name]


def get_path_chose_file():
    options = QFileDialog.Options()
    options |= QFileDialog.DontUseNativeDialog

    file_filter = "Word Files (*.doc *.docx);;PDF Files (*.pdf)"

    file_path, _ = QFileDialog.getOpenFileName(None, "Select File", "", file_filter, options=options)

    return file_path


def set_filename():
    ui.URL_lineEdit.setText(get_path_chose_file())


def url_is_correct(path):
    if os.path.isfile(path):
        return True
    if requests.head(path).status_code == 200:
        return True
    return False


def read_parameters_display():
    diagonal = ui.Diagonal_LineEdit.text()
    width_px = ui.Distance_LineEdit_2.text()
    height_px = ui.Distance_LineEdit_3.text()
    distance = ui.Distance_LineEdit.text()
    return diagonal, width_px, height_px, distance


def read_path_file_or_site():
    return ui.URL_lineEdit.text()


def showMessage(message):
    ui.add_text_to_process_textedit(message)


def start_analyze():
    path = read_path_file_or_site()

    if path == "":
        msg = QtWidgets.QMessageBox()
        msg.setWindowTitle("Попередження")
        msg.setText("Відсутній шлях до файлу чи веб-сторінки!")
        msg.exec_()
        return

    if url_is_correct(path) is not True:
        msg = QtWidgets.QMessageBox()
        msg.setWindowTitle("Попередження")
        msg.setText("Сайт або файл не знайдено!")
        msg.exec_()
        return

    model = Model()
    diagonal, width_px, height_px, distance = read_parameters_display()

    if not diagonal or not width_px or not height_px or not distance:
        msg = QtWidgets.QMessageBox()
        msg.setWindowTitle("Помилка")
        msg.setText("Будь ласка, заповніть усі поля!")
        msg.exec_()
        return

    try:
        diagonal = int(diagonal)
        width_px = int(width_px)
        height_px = int(height_px)
        distance = int(distance)
    except ValueError:
        msg = QtWidgets.QMessageBox()
        msg.setWindowTitle("Помилка")
        msg.setText("Усі значення повинні бути числами!")
        msg.exec_()
        return

    if distance == 0 or diagonal == 0 or width_px == 0 or height_px == 0:
        msg = QtWidgets.QMessageBox()
        msg.setWindowTitle("Попередження")
        msg.setText("Значення полів не можуть бути рівні нулю!")
        msg.exec_()
        return

    model.set_ppi(width_px, height_px, diagonal)
    ui.DPI_LineEdit.setText(str(round(model.PPI,0)))
    model.set_distance_to_display(int(distance))

    if "http" in path or "https" in path:
        model.set_path(path)
        model.read_text_from_site(path)

    else:
        if "docx" in path or "doc" in path:
            model.convert_word_to_pdf(path)
        else:
            model.set_path(path)
        model.read_text_from_pdf()

    showMessage("Починається обробка...")

    words_spans = model.get_text_list_spans()

    for word in words_spans:
        showMessage(f"{repr(word.text_span)}")

    showMessage("Loading frequency dictionary...")

    workbook = get_frequency_dictionary('data/wordFrequency.xlsx')
    sheet = get_sheet_workbook(workbook, "4 forms (219k)")

    if not sheet:
        showMessage("Sheet not found!")
        exit()

    column_b = get_column_by_sheet(sheet, 'B')
    biggest_value_freq = get_biggest_frequency(sheet)

    rest_letters = 0
    state = "updated"

    for word in words_spans:
        if word.text_span.isalpha():
            showMessage(f"Наступне слово : {word.text_span}")
            showMessage(f"Властивості шрифту : {word.font_span}, {int(word.size)}")

            if rest_letters > 3:
                rest_letters = 3

            index_chose = 0

            dict_probability = model.calculate_probability_landing(word.text_span, rest_letters)

            showMessage(f"Розрахунок індекса слова : <{word.text_span}>")

            if state == "updated":
                index_chose = model.calculate_final_pos_fixation(dict_probability)
                showMessage(f"Індекс <{index_chose}> було обрано!")

            if state == "2 symbols after word":
                index_chose = 0
                state = "updated"

            if word.distance_to_next_span > 0:
                index_chose = len(word.text_span) - 1
                showMessage("Кінець цього слова...")

            if word.is_last_in_line:
                index_chose = len(word.text_span) - 1
                showMessage("Кінець рядку...")

            if index_chose == len(word.text_span):
                showMessage("Слово було пропущене! Приземлення на наступний символ!")
                rest_letters = 0

            elif index_chose > len(word.text_span):
                showMessage("Слово пропущене! Приземлення на 2 символа після слова!")
                rest_letters = 0
                state = "2 symbols after word"

            else:
                showMessage(f"Фіксація в слові <{word.text_span}> на слові {word.text_span[index_chose]}!")
                rest_letters = len(word.text_span) - index_chose
                prob_refix = model.calculate_probability_refixation(word.text_span, index_chose)
                showMessage(f"Вірогідність рефіксації = {round(prob_refix, 3)}")

                if model.should_refixate(prob_refix):
                    showMessage("------------------Необхідна рефіксація------------------")

                    time_refix = model.make_refixation(word.text_span, index_chose + 1)
                    showMessage(f"Час затримки рефіксації = {round(time_refix, 3)}")

                    time_refix_sd = model.calculate_sd(time_refix)

                    model.increase_general_time(time_refix)
                    model.increase_general_time_sd(time_refix_sd)

                else:
                    showMessage("Рефіксація не потрібна...")

                freq = find_freq_for_word(sheet, column_b, word.text_span)
                showMessage("Обчислення часу лексичної ідентифікації...")

                time_word_reading = model.calculate_time_reading(word.text_span, index_chose + 1, biggest_value_freq,
                                                                 freq)
                showMessage(f"Для слова <{word.text_span}> необхідний час читання = {round(time_word_reading, 3)}")

                time_to_read_sd = model.calculate_sd(time_word_reading)
                dispersion = model.calculate_normal_distribution(time_word_reading, time_to_read_sd)

                model.increase_general_time(dispersion)

                showMessage("Виконання сакади...")

            model.add_average_latency_time()
            model.add_standard_deviation_latency_time()

        else:
            rest_letters += len(word.text_span)

        if word.distance_to_next_span > 0:
            time_saccade = model.calculate_time_saccade(word.distance_to_next_span)
            showMessage(f"Перехід на наступний блок... Необхідно часу {time_saccade} ms\n")
            model.increase_general_time(time_saccade)

    show_time_to_read(model.get_sum_time_reading(), model.get_sum_standard_deviation())
    showMessage("Аналіз завершено.")


if __name__ == "__main__":

    import sys
    app = QtWidgets.QApplication(sys.argv)
    MainWindow = QtWidgets.QMainWindow()
    ui = Ui_MainWindow()
    ui.setupUi(MainWindow)
    MainWindow.show()

    ui.Start_btn.clicked.connect(start_analyze)
    ui.ChooseFile_btn.clicked.connect(set_filename)

    sys.exit(app.exec_())
