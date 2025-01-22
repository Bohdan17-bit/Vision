from openpyxl import load_workbook
from PyQt5.QtCore import QThread, pyqtSignal
import re


class FrequencyDictionary(QThread):
    finished = pyqtSignal()

    def __init__(self):
        super().__init__()
        self.workbook = None
        self.sheet = None
        self.column = None

        self.ones = [
            "zero", "one", "two", "three", "four", "five", "six", "seven", "eight", "nine"
        ]
        self.teens = [
            "ten", "eleven", "twelve", "thirteen", "fourteen", "fifteen",
            "sixteen", "seventeen", "eighteen", "nineteen"
        ]
        self.tens = [
            "", "", "twenty", "thirty", "forty", "fifty",
            "sixty", "seventy", "eighty", "ninety"
        ]
        self.thousands = ["", "thousand"]

        self.abbreviations = {
            "etc.": "et cetera",
            "etc": "et cetera",
            "eg": "for example",
            "i.e.": "that is",
            "i.e": "that is",
            "vs.": "versus",
            "vs": "versus",
            "a.m.": "am",
            "am": "am",
            "pm": "pm",
            "p.m.": "pm",
            "approx.": "approximately",
            "Dr.": "Doctor",
            "Mr.": "Mister",
            "Mrs.": "Mistress",
            "Ph.D.": "Doctor of Philosophy",
            "B.Sc.": "Bachelor of Science",
            "M.Sc.": "Master of Science",
            "Inc.": "Incorporated",
            "Ltd.": "Limited",
            "St.": "Saint",
            "Jr.": "Junior",
            "Sr.": "Senior",
        }

        self.contraction_map = {
            "wasn't": "was not", "wasnt": "was not", "was`nt": "was not",
            "weren't": "were not", "werent": "were not", "weren`t": "were not",
            "i'm": "I am", "i`m": "I am",
            "you've": "you have", "youve": "you have",
            "we've": "we have", "weve": "we have",
            "they've": "they have", "theyve": "they have",
            "he's": "he is", "hes": "he is", "he`s": "he is",
            "she's": "she is", "shes": "she is", "she`s": "she is",
            "it's": "it is", "its": "it is", "it`s": "it is",
            "i'd": "I would", "i`d": "I would",
            "you'd": "you would", "you`d": "you would",
            "we'd": "we would", "we`d": "we would",
            "they'd": "they would", "they`d": "they would",
            "he'd": "he would", "he`d": "he would",
            "she'd": "she would", "she`d": "she would",
            "it'd": "it would", "it`d": "it would",
            "i'll": "I will", "i`ll": "I will",
            "you'll": "you will", "you`ll": "you will",
            "we'll": "we will", "we`ll": "we will",
            "they'll": "they will", "they`ll": "they will",
            "he'll": "he will", "he`ll": "he will",
            "she'll": "she will", "she`ll": "she will",
            "it'll": "it will", "it`ll": "it will",
            "don't": "do not", "dont": "do not", "don`t": "do not",
            "doesn't": "does not", "doesnt": "does not", "doesn`t": "does not",
            "didn't": "did not", "didnt": "did not", "didn`t": "did not",
            "can't": "cannot", "cant": "cannot", "can`t": "cannot",
            "won't": "will not", "wont": "will not", "won`t": "will not",
            "wouldn't": "would not", "wouldnt": "would not", "wouldn`t": "would not",
            "shouldn't": "should not", "shouldnt": "should not", "shouldn`t": "should not",
            "couldn't": "could not", "couldnt": "could not", "couldn`t": "could not",
            "mightn't": "might not", "mightnt": "might not", "mightn`t": "might not",
            "mustn't": "must not", "mustnt": "must not", "mustn`t": "must not",
            "who's": "who is", "who`'s": "who is",  # Для питання "who is"
            "here's": "here is", "there's": "there is",  # Для "here is" і "there is"
            "that's": "that is", "this's": "this is",  # "that is", "this is"
            "isn't": "is not", "isnt": "is not", "isn`t": "is not",  # "is not"
            "aren't": "are not", "arent": "are not", "aren`t": "are not",  # "are not"
            "hadn't": "had not", "hadnt": "had not", "hadn`t": "had not",  # "had not"
            "let's": "let us",  # "let us"
            "what's": "what is",  # "what is"
            "how's": "how is",  # "how is",
            "etc": "et cetera", "etc.": "et cetera",
            "Etc": "et cetera",
            "I’m": "I am",
            "I'm": "I am",
            "I`m": "I am",
            "i’m": "I am"
        }

    def run(self):
        self.load_dictionary()

    def load_dictionary(self):
        self.workbook = self.get_frequency_dictionary('data/wordFrequency.xlsx')
        self.sheet = self.get_sheet_workbook(self.workbook, "4 forms (219k)")
        self.column = self.get_column_by_sheet('B')
        self.finished.emit()

    def get_frequency_dictionary(self, path):
        return load_workbook(path)

    def get_sheet_workbook(self, workbook, sheet):
        list_sheets = workbook.sheetnames
        if sheet in list_sheets:
            return workbook[sheet]
        else:
            return None

    def find_freq_for_word(self, sheet, column, word):
        for cell in column:
            if word == cell.value and cell.row != 1:
                return sheet.cell(row=cell.row, column=15).value
        return 0

    def get_biggest_frequency(self):
        return self.sheet.cell(row=2, column=15).value

    def get_column_by_sheet(self, col):
        return self.sheet[col]

    def parse_hundreds(self, num):
        if num == 0:
            return ""
        if num < 10:
            return self.ones[num]
        elif num < 20:
            return self.teens[num - 10]
        else:
            result = self.tens[num // 10]
            if num % 10 != 0:
                result += " " + self.ones[num % 10]
            return result

    def number_to_words(self, n):
        if n == 0 or n == '0':
            return "zero"

        str_n = str(n)
        words = []

        # Зберігаємо лідируючі нулі
        leading_zeros = len(str_n) - len(str_n.lstrip('0'))
        words.extend(['zero'] * leading_zeros)

        # Обробка решти числа
        remaining_number = int(str_n.lstrip('0')) if str_n.lstrip('0') else 0
        if remaining_number <= 9999:
            if remaining_number >= 1000:
                words.append(self.ones[remaining_number // 1000] + " " + self.thousands[1])
                remaining_number %= 1000
            if remaining_number >= 100:
                words.append(self.ones[remaining_number // 100] + " hundred")
                remaining_number %= 100
            if remaining_number > 0:
                words.append(self.parse_hundreds(remaining_number))
        else:
            for digit in str(remaining_number):
                words.append(self.ones[int(digit)])

        return " ".join(words).strip()

    def split_mixed_word(self, text):
        segments = re.findall(r"[a-zA-Z]+(?:['`’][a-zA-Z]+)?|\d+", text)
        return segments

    def mixed_word_to_words(self, text):
        segments = self.split_mixed_word(text)
        result = []

        for segment in segments:
            lower_segment = segment.lower()

            if lower_segment in self.abbreviations:
                pronounced_segment = self.abbreviations[lower_segment]
                result.append(pronounced_segment)

            # Обробка скорочень
            elif lower_segment in self.contraction_map:
                result.append(self.contraction_map[lower_segment])

            # Обробка апострофів та лапок
            elif "`" in segment or "’" in segment or "'" in segment:
                parts = re.split(r"['`’]", segment)
                for part in parts:
                    if part:
                        result.append(part)

            # Обробка чисел
            elif segment.isdigit():
                if all(char == '0' for char in segment):
                    result.append(" ".join("zero" for _ in segment))
                else:
                    # Використовуємо метод, який враховує лідируючі нулі
                    result.append(self.number_to_words_with_leading_zeros(segment))

            # Обробка змішаних слів
            elif any(char.isdigit() for char in segment):
                result.append(self.convert_digit_to_pronunciation(segment))


            # Інші випадки
            else:
                result.append(segment)

        return " ".join(result)

    def convert_digit_to_pronunciation(self, segment):
        pronunciation = ""
        for char in segment:
            if char.isdigit():
                pronunciation += f"{self.ones[int(char)]} "
            else:
                pronunciation += char
        return pronunciation.strip()

    def number_to_words_with_leading_zeros(self, n):
        str_n = str(n)
        leading_zeros = len(str_n) - len(str_n.lstrip('0'))
        words = ['zero'] * leading_zeros

        if str_n.lstrip('0'):
            words.append(self.number_to_words(int(str_n.lstrip('0'))))

        return " ".join(words)